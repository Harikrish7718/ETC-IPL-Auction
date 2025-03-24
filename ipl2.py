import pandas as pd
import re
from openpyxl import load_workbook

# File path and reading lines from ipl.txt
file_path = "ipl.txt"  
with open(file_path, "r", encoding="utf-8") as file:
    lines = [line.strip() for line in file if line.strip()]  

cleaned_lines = []
start_processing = False  

# Process lines based on the starting criteria
for line in lines:
    if line.startswith("1"):  
        start_processing = True

    if start_processing:
        cleaned_lines.append(line)

# Extract player points from the file
player_points = {}
i = 0  

while i < len(lines) - 3:  
    if lines[i].isdigit():  
        player_name = lines[i + 1].strip()  
        team_name = lines[i + 2].strip()  
        stats_line = lines[i + 3].strip()  

        stats_parts = stats_line.split()
        if stats_parts and stats_parts[0].replace(".", "", 1).isdigit():  
            points = float(stats_parts[0])  
            player_points[player_name] = points
        else:
            print(f"Skipping invalid data for player: {player_name}")
    i += 1  

# Load the existing Excel file (without creating a new sheet)
excel_file = "ETC IPL Auction '25  Mastersheet 22_Mar.xlsx"
try:
    # Open the existing workbook
    wb = load_workbook(excel_file)
    
    # Access the 'Players' sheet
    ws = wb["Players"]  

    # Debugging: Print all player names in the Excel sheet to see how they look
    print("Player Names in Excel Sheet:")
    for row in range(2, ws.max_row + 1):  # Start from row 2 (assuming row 1 is header)
        player_name_in_excel = ws.cell(row=row, column=5).value  # Player Name is in column E (5th column)
        print(f"Excel: {player_name_in_excel}")

    # Loop through the rows of the 'Players' sheet
    for row in range(2, ws.max_row + 1):  # Start from row 2 (assuming row 1 is header)
        player_name = ws.cell(row=row, column=5).value.strip()  # Player Name from Excel (Column E)
        
        # Check if the player name exists in player_points
        if player_name in player_points:
            ws.cell(row=row, column=6, value=player_points[player_name])  # Update column F (6th column)
        else:
            print(f"Player not found in data: {player_name}")  # Debug print to track if there's a mismatch

    # Save the updated file with values only
    wb.save(excel_file)  # This saves over the original file, preserving other sheets

    print(f"Excel file updated successfully: {excel_file}")

except Exception as e:
    print(f"Error occurred: {e}")
